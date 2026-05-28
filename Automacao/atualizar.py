import openpyxl, json, re, os, shutil, subprocess, sys
from collections import defaultdict
from datetime import datetime

# ── Caminhos ──────────────────────────────────────────────────────────────────
BASE       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL      = os.path.join(BASE, '2025 - Financeiro Motiva - Controle Lorena.v2.xlsx')
DASHBOARD  = os.path.join(BASE, 'dashboard_motiva_v2.html')
INDEX      = os.path.join(BASE, 'index.html')   # GitHub Pages e fallback
GIT_REPO   = BASE

# Token armazenado no git config local — nao incluir no codigo

# ── Helpers ───────────────────────────────────────────────────────────────────
def safe_float(v):
    try: return float(v) if v and str(v).strip() not in ('#N/A','#REF!','#VALUE!','None','') else 0.0
    except: return 0.0

def month_key(val):
    if not val: return None
    if hasattr(val, 'strftime'): return val.strftime('%Y_%m')
    s = str(val).strip()
    return s if len(s)==7 and '_' in s else None

def norm_vendor(v):
    if not v: return None
    m = {'karine':'Karine','carol':'Carol','andre':'Andre','andre':'André',
         'lysandra':'Lysandra','vanessa':'Vanessa','pedro':'Pedro'}
    return m.get(str(v).strip().lower(), str(v).strip())

def replace_const(html, name, vals, strings=False):
    if strings:
        vs = ','.join(f"'{v}'" for v in vals)
    else:
        vs = ','.join(str(int(v)) if float(v)==int(float(v)) else str(round(float(v),2)) for v in vals)
    pat = rf'(const\s+{re.escape(name)}\s*=\s*\[)[^\]]*(\])'
    new, n = re.subn(pat, rf'\g<1>{vs}\g<2>', html)
    print(f"  {'OK' if n else 'XX'} {name}: {len(vals)} valores")
    return new

def validar_html(html, label=''):
    ok = True
    if not html.rstrip().endswith('</html>'):
        print(f"  AVISO {label}: arquivo NAO termina com </html>!")
        ok = False
    if html.count('</script>') < 2:
        print(f"  AVISO {label}: menos de 2 blocos script!")
        ok = False
    if len(html) < 200000:
        print(f"  AVISO {label}: arquivo pequeno ({len(html)//1024}KB)!")
        ok = False
    return ok

# ── 1. ACOMPANHAMENTO ─────────────────────────────────────────────────────────
print("\nLendo Acompanhamento...")
wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
ws = wb['Acompanhamento']
hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
N = sum(1 for i in range(7, len(hdr) if hdr else 30) if hdr[i] is not None)
if N < 12: N = 17
print(f"  Meses: {N}")

NOMES = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
LABELS = []
y, mo = 2025, 1
for _ in range(N):
    LABELS.append(f"{NOMES[mo-1]}/{str(y)[2:]}"); mo+=1
    if mo>12: mo,y=1,y+1

CS, CE = 7, 7+N
rows = {i: list(row) for i,row in enumerate(ws.iter_rows(min_row=1,max_row=200,values_only=True),1)}
wb.close()

def rs(rn):
    row = rows.get(rn,[])
    return [int(round(safe_float(row[j] if j<len(row) else None))) for j in range(CS,CE)]

gmvVT=rs(11); gmvPAT=rs(15); gmvTotal=rs(19)
recVT=rs(33); recPAT=rs(37)
commKarine=rs(132); commCarol=rs(134); commAndre=rs(138)
commLysandra=rs(152); commVanessa=rs(154); commPedro=rs(156)
commTotal=[commKarine[i]+commCarol[i]+commAndre[i]+commLysandra[i]+commVanessa[i]+commPedro[i] for i in range(N)]
print(f"  GMV VT: R$ {sum(gmvVT):,.0f}  |  Comissao: R$ {sum(commTotal):,.0f}")

# ── 2. GMV POR CLIENTE ────────────────────────────────────────────────────────
print("\nLendo GMV por Cliente...")
wb2 = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
ws2 = wb2['GMV por Cliente']
SEG_IDS = {'Diamante':0,'Ouro':1,'Prata':2,'Bronze-P':3,'Bronze-PP':4}

def fmt_cnpj(v):
    if not v: return ''
    s=str(v).replace('.0','').strip().zfill(14)
    try: return f"{s[:2]}.{s[2:5]}.{s[5:8]}/{s[8:12]}-{s[12:14]}"
    except: return str(v)

def fmt_date(v):
    if not v: return ''
    try:
        if hasattr(v,'strftime'): return v.strftime('%m/%y')
        parts=str(v)[:7].split('-')
        return f"{parts[1]}/{parts[0][2:]}" if len(parts)>=2 else ''
    except: return ''

clients=[]
for row in ws2.iter_rows(min_row=3,values_only=True):
    if not row[1]: continue
    gmv_m=[round(safe_float(row[j] if j<len(row) else None),1) for j in range(CS,CE)]
    if sum(gmv_m)==0: continue
    clients.append({'empresa':str(row[1]).strip(),'cnpj':fmt_cnpj(row[2]),
                    'seg':str(row[5]).strip() if row[5] else '','cad':fmt_date(row[4]),'gmv':gmv_m})
wb2.close()
clients.sort(key=lambda c:sum(c['gmv']),reverse=True)
print(f"  {len(clients)} clientes")

cd_parts=[]
for c in clients:
    gmv_s=','.join(str(v) for v in c['gmv'])
    cd_parts.append(f"[{json.dumps(c['cnpj'])},{json.dumps(c['empresa'],ensure_ascii=False)},{SEG_IDS.get(c['seg'],5)},{json.dumps(c['cad'])},0,{gmv_s},0]")
CD_JS='const _CD=[\n'+',\n'.join(cd_parts)+'\n];'

# ── 3. PEDIDOS POR VENDEDOR + TAXAS ───────────────────────────────────────────
print("\nLendo pedidos por vendedor e taxas...")
MK=[]; yy,mm=2025,1
for _ in range(N):
    MK.append(f"{yy}_{mm:02d}"); mm+=1
    if mm>12: mm,yy=1,yy+1
MIDX={k:i for i,k in enumerate(MK)}

vd=defaultdict(lambda:defaultdict(lambda:{'cnpj':'','pm':defaultdict(lambda:{'g':0,'r':0,'c':0})}))
pcc_adm=defaultdict(float); pcc_suc=defaultdict(float)
pcc_ent=defaultdict(float); pcc_nf=defaultdict(float)
pe_adm=defaultdict(float);  pe_suc=defaultdict(float)

wb3=openpyxl.load_workbook(EXCEL,data_only=True,read_only=True)
for row in wb3['Pedido de carga Completo'].iter_rows(min_row=2,values_only=True):
    mes=month_key(row[8] if len(row)>8 else None)
    vend=norm_vendor(row[23] if len(row)>23 else None)
    emp=str(row[4]).strip() if len(row)>4 and row[4] else None
    if not mes or not vend or not emp or mes not in MIDX: continue
    d=vd[vend][emp]
    if len(row)>1 and row[1]: d['cnpj']=str(row[1]).strip()
    p=d['pm'][mes]
    p['g']+=safe_float(row[11] if len(row)>11 else None)
    p['r']+=safe_float(row[15] if len(row)>15 else None)+safe_float(row[16] if len(row)>16 else None)+safe_float(row[18] if len(row)>18 else None)
    p['c']+=safe_float(row[26] if len(row)>26 else None)
    pcc_adm[mes]+=safe_float(row[15] if len(row)>15 else None)
    pcc_suc[mes]+=safe_float(row[16] if len(row)>16 else None)
    pcc_ent[mes]+=safe_float(row[18] if len(row)>18 else None)
    pcc_nf[mes] +=safe_float(row[19] if len(row)>19 else None)

for row in wb3['Pedido externo'].iter_rows(min_row=2,values_only=True):
    mes=month_key(row[4] if len(row)>4 else None)
    vend=norm_vendor(row[20] if len(row)>20 else None)
    emp=str(row[5]).strip() if len(row)>5 and row[5] else None
    if not mes or not vend or not emp or mes not in MIDX: continue
    d=vd[vend][emp]
    if not d['cnpj'] and len(row)>17 and row[17]: d['cnpj']=str(row[17]).strip()
    p=d['pm'][mes]
    p['g']+=safe_float(row[7] if len(row)>7 else None)
    p['r']+=safe_float(row[15] if len(row)>15 else None)
    p['c']+=safe_float(row[23] if len(row)>23 else None)
    pe_adm[mes]+=safe_float(row[13] if len(row)>13 else None)
    pe_suc[mes]+=safe_float(row[14] if len(row)>14 else None)
wb3.close()

vtTxAdm    =[int(round(pcc_adm[mk]+pe_adm[mk])) for mk in MK]
vtTxSucesso=[int(round(pcc_suc[mk]+pe_suc[mk])) for mk in MK]
vtTxEntrega=[int(round(pcc_ent[mk]))             for mk in MK]
nfPedCarga =[int(round(pcc_nf[mk]))              for mk in MK]
print(f"  Tx Adm: R$ {sum(vtTxAdm):,.0f} | Tx Sucesso: R$ {sum(vtTxSucesso):,.0f} | NF: R$ {sum(nfPedCarga):,.0f}")

VENDORS=['Karine','André','Carol','Lysandra','Vanessa','Pedro']
vd_out={}
for vend in VENDORS:
    items=[]
    for emp,info in vd.get(vend,{}).items():
        ga=[round(info['pm'].get(mk,{}).get('g',0),2) for mk in MK]
        ra=[round(info['pm'].get(mk,{}).get('r',0),2) for mk in MK]
        ca=[round(info['pm'].get(mk,{}).get('c',0),2) for mk in MK]
        if sum(ca)==0: continue
        items.append((sum(ga),emp,info['cnpj'],ga,ra,ca))
    items.sort(reverse=True)
    vd_out[vend]=[(e,c,ga,ra,ca) for _,e,c,ga,ra,ca in items]
    print(f"  {vend}: {len(vd_out[vend])} empresas")

def ajson(arr): return '['+','.join(str(round(v,2)) for v in arr)+']'
vd_parts=[]
for vend in VENDORS:
    ep=[f"[{json.dumps(e,ensure_ascii=False)},{json.dumps(c)},{ajson(ga)},{ajson(ra)},{ajson(ca)}]"
        for e,c,ga,ra,ca in vd_out.get(vend,[])]
    vd_parts.append(f"{json.dumps(vend,ensure_ascii=False)}:[{','.join(ep)}]")
VD_JS='const VD={'+','.join(vd_parts)+'};'

# ── 4. ATUALIZAR HTML ─────────────────────────────────────────────────────────
print("\nAtualizando HTML...")
with open(DASHBOARD,'r',encoding='utf-8') as f: html=f.read()
orig_kb=len(html)//1024

if not validar_html(html,'arquivo original'):
    print("\nERRO: Arquivo original corrompido. Abortando.")
    sys.exit(1)

html=replace_const(html,'months',      LABELS,      strings=True)
html=replace_const(html,'gmvVT',       gmvVT)
html=replace_const(html,'gmvPAT',      gmvPAT)
html=replace_const(html,'gmvTotal',    gmvTotal)
html=replace_const(html,'recVT',       recVT)
html=replace_const(html,'recPAT',      recPAT)
html=replace_const(html,'vtTxAdm',     vtTxAdm)
html=replace_const(html,'vtTxSucesso', vtTxSucesso)
html=replace_const(html,'vtTxEntrega', vtTxEntrega)
html=replace_const(html,'nfPedCarga',  nfPedCarga)
html=replace_const(html,'commTotal',   commTotal)
html=replace_const(html,'commKarine',  commKarine)
html=replace_const(html,'commAndre',   commAndre)
html=replace_const(html,'commCarol',   commCarol)
html=replace_const(html,'commLysandra',commLysandra)
html=replace_const(html,'commVanessa', commVanessa)
html=replace_const(html,'commPedro',   commPedro)

html,n=re.subn(r'const VD=\{.*?\};',VD_JS,html,flags=re.DOTALL)
print(f"  {'OK' if n else 'XX'} VD: {sum(len(v) for v in vd_out.values())} empresas")
html,n=re.subn(r'const _CD=\[.*?\];',CD_JS,html,flags=re.DOTALL)
print(f"  {'OK' if n else 'XX'} _CD: {len(clients)} clientes")

html=re.sub(r'emandamento',f'em andamento',html)
html=re.sub(r'em\s+andamento.*?(?=<)',f'em andamento',html)
html=re.sub(r'andamento[^<]*','andamento',html)
html=re.sub(r'em andamento','em andamento',html)
html=re.sub(r'[⚡]\s*\w+/\d+\s*em andamento',f'⚡ {LABELS[-1]} em andamento',html)
html=re.sub(r'Dashboard Financeiro\s*\xb7\s*\w+/\d+\s*–\s*\w+/\d+',
            f'Dashboard Financeiro \xb7 {LABELS[0]} – {LABELS[-1]}',html)
print(f"  OK Badge: {LABELS[0]} a {LABELS[-1]}")

if not validar_html(html,'arquivo atualizado'):
    print("\nERRO: Resultado corrompido. Abortando.")
    sys.exit(1)

with open(DASHBOARD,'w',encoding='utf-8') as f: f.write(html)
shutil.copy2(DASHBOARD,INDEX)
print(f"\nHTML salvo! {orig_kb}KB -> {len(html)//1024}KB | index.html copiado")

# ── 5. PUSH PARA GITHUB ───────────────────────────────────────────────────────
print("\nPublicando no GitHub...")
data_hoje=datetime.now().strftime('%d/%m/%Y %H:%M')
try:
    def git(args):
        r=subprocess.run(['git','-C',GIT_REPO]+args,capture_output=True,text=True)
        if r.returncode!=0 and r.stderr: print(f"  git {args[0]}: {r.stderr.strip()}")
        return r
    git(['config','user.email','vitor.leite@usekim.com.br'])
    git(['config','user.name','Vitor'])
    git(['add','dashboard_motiva_v2.html','index.html'])
    commit=git(['commit','-m',f'Auto-update {data_hoje} -- {LABELS[-1]} em andamento'])
    if 'nothing to commit' in commit.stdout:
        print("  Sem alteracoes desde ultimo commit.")
    else:
        push=git(['push','origin','main'])
        if push.returncode==0:
            print(f"  Publicado! -> https://vitorabl.github.io/motiva-dashboard/")
        else:
            print(f"  Push falhou: {push.stderr.strip()}")
except Exception as e:
    print(f"  Erro no git: {e}")
    print("  Dashboard salvo localmente. Faca push manual se necessario.")

print(f"\nProcesso concluido em {datetime.now().strftime('%H:%M:%S')}")
